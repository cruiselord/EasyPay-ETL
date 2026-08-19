"""
Writer module for the NIBSS ETL pipeline.

Builds a 5-sheet Excel workbook from the cleaned DataFrames produced
by ``reader.load_all_data()``.

Sheets
------
1. Summary    — pivot-table of counts, success rate and total value per session
2. EasyPay    — all Easy Pay rows (credit / debit / name_enquiry) in a filterable table
3. DirectDebit — all Direct Debit rows (credit / debit) in a filterable table
4. Exceptions — failed rows from both sources, sorted by amount descending
5. Run_Log    — pipeline metadata / file manifest

Uses ``openpyxl`` directly so we can control table styles, AutoFilter,
and conditional row colouring with fine granularity.
"""

from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule


# ── Style constants ──────────────────────────────────────────────────────────

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TOP_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="medium"),
    bottom=Side(style="thin"),
)

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

# Priority column order for data sheets.
_PRIORITY_COLUMNS = [
    "session", "transaction_type", "status", "amount",
    "transaction_id", "payment_reference", "narration",
    "response_status", "message", "nip_response_code",
]

SESSION_LABEL_MAP = {
    "0559": "5:59 AM",
    "1159": "11:59 AM",
    "1759": "5:59 PM",
    "2359": "11:59 PM",
}

NIP_CODE_MAP = {
    "00": "Successful",
    "91": "Routing error / Beneficiary bank unavailable",
    "96": "System malfunction",
    "97": "Timeout waiting for response",
}

# Bank short-name → full-name mapping (from beneficiary_bvn column).
# Some codes are NIBSS-assigned abbreviations; uncertain entries noted.
BANK_NAME_MAP = {
    "ABP": "Access Bank",
    "DBP": "Diamond Bank (Legacy)",
    "ECO": "Ecobank Nigeria",
    "FAIRM": "FBNQuest Merchant Bank",
    "FBN": "First Bank of Nigeria",
    "FBP": "First Bank of Nigeria",
    "FCM": "FCMB",
    "GTB": "GTBank (GTCO)",
    "": "Not Specified",
    "IBT": "Ibeto MFB",
    "KDB": "Kuda Bank",
    "MONIEPT": "Moniepoint MFB",
    "OPY": "OPay",
    "PLM": "Polaris Bank",
    "PRLLX": "Parallel MFB",
    "PRMTRST": "Promise Trust MFB",
    "SBP": "Stanbic IBTC Bank",
    "SKY": "Skye Bank (Legacy)",
    "TAJ": "Taj Bank",
    "UBA": "UBA",
    "UBN": "Union Bank",
    "UMB": "Unity Bank",
    "VDM": "VFD MFB",
    "WMA": "Wema Bank",
    "ZIB": "Zenith Bank",
    "9PSB": "9 Payment Service Bank",
}

# Channel-code descriptions
CHANNEL_DESC = {
    "2": "NIBSS NIP Channel",
    "100004": "OPay (originating)",
    "000012": "Institution 000012",
    "000016": "First Bank (originating)",
}


# ── Shared helpers ───────────────────────────────────────────────────────────

def _apply_header_style(ws, max_col: int):
    """Style the first row as a blue table header."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_red_to_failed(ws, status_col: str, max_row: int):
    """
    Red-fill every row where the *status_col* value is ``"failed"``
    (the filename-derived classification, not the original CSV status).
    """
    for row in range(2, max_row + 1):
        cell = ws[f"{status_col}{row}"]
        if cell.value and str(cell.value).strip().lower() == "failed":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col_idx).fill = RED_FILL
                ws.cell(row=row, column=col_idx).font = RED_FONT


def _auto_width(ws):
    """Auto-adjust column widths (capped at 60 so long strings don't balloon)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def _find_col_letter(ws, header_value: str) -> str | None:
    """
    Return the Excel column letter for the first column whose header
    matches *header_value* (case-insensitive).  Returns ``None`` if
    not found.
    """
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value and str(cell.value).strip().lower() == header_value.lower():
            return cell.column_letter
    return None


def _apply_currency_format(ws, col_letter: str, max_row: int):
    """Apply ``"₦"#,##0.00`` number format to every data cell in a column."""
    for row in range(2, max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        if cell.value is not None:
            cell.number_format = '"₦"#,##0.00'


# ── Column ordering ──────────────────────────────────────────────────────────

def _reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reorder columns: priority columns first (session, type, status, amount,
    txn_id, ref, narration, response, message, nip), then remaining columns
    alphabetically.  Columns absent from the DataFrame are silently skipped.
    """
    if df.empty:
        return df
    existing = [c for c in _PRIORITY_COLUMNS if c in df.columns]
    remaining = sorted([c for c in df.columns if c not in _PRIORITY_COLUMNS])
    return df[existing + remaining]


# ── Dead-column audit ────────────────────────────────────────────────────────

def _remove_dead_columns(df: pd.DataFrame, sheet_name: str, run_log: list) -> pd.DataFrame:
    """
    Drop columns that are 100 % NaN/empty across all rows and record the
    removal in *run_log*.  Only non-empty DataFrames are inspected.
    """
    if df.empty:
        return df
    dead = [col for col in df.columns if df[col].isna().all()]
    if dead:
        df = df.drop(columns=dead)
        run_log.append({
            "session": "ALL",
            "source": sheet_name,
            "file": f"(dropped {len(dead)} dead columns)",
            "found": True,
            "rows": 0,
            "size_kb": None,
            "modified": None,
            "detail": f"100 % null columns removed: {dead}",
        })
    return df


# ── Table writer ─────────────────────────────────────────────────────────────

def _df_to_table(ws, df: pd.DataFrame, table_name: str):
    """
    Write *df* into *ws* starting at cell A1 and register it as an
    Excel Table with AutoFilter enabled.
    """
    if df.empty:
        ws.cell(row=1, column=1, value="(no data)")
        return

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    nrows = len(df) + 1
    ncols = len(df.columns)
    if nrows > 1 and ncols > 0:
        col_letter = chr(64 + ncols) if ncols <= 26 else "Z"
        table_ref = f"A1:{col_letter}{nrows}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TABLE_STYLE
        ws.add_table(table)


# ── Sheet builders ───────────────────────────────────────────────────────────

def _bank_name(code: str) -> str:
    """Resolve a bank short-code to full name. Returns the code itself if unknown."""
    return BANK_NAME_MAP.get(code, code)


def _compute_kpis(combined: pd.DataFrame) -> pd.DataFrame:
    """Headline KPIs returned as a 1-row DataFrame."""
    total = len(combined)
    total_val = combined["amount"].fillna(0).sum()
    ok = combined[combined["status"] == "successful"]
    fail = combined[combined["status"] == "failed"]
    ok_val = ok["amount"].fillna(0).sum()
    fail_val = fail["amount"].fillna(0).sum()
    sr = (len(ok) / total * 100) if total else 0
    vsr = (ok_val / total_val * 100) if total_val else 0
    banks = int(combined["beneficiary_bvn"].nunique()) if "beneficiary_bvn" in combined.columns else 0
    return pd.DataFrame([{
        "Total Txns": total, "Successful": len(ok), "Failed": len(fail),
        "Rate": f"{sr:.1f}%", "Total Value": round(total_val, 2),
        "At Risk": round(fail_val, 2), "Val Rate": f"{vsr:.1f}%",
        "Avg Value": round(total_val / total, 2) if total else 0, "Banks": banks,
    }])


def _compute_session_overview(combined: pd.DataFrame) -> pd.DataFrame:
    """Per-session pivot with count + value splits."""
    c = combined.copy()
    c["amount"] = c["amount"].fillna(0)
    pivot = c.pivot_table(index="session", columns=["transaction_type", "status"],
                          values="amount", aggfunc="count", fill_value=0)
    pivot.columns = [f"{t}_{s}" for t, s in pivot.columns]
    agg = c.groupby("session").agg(txns=("amount","count"), tv=("amount","sum"),
                                   fv=("amount",lambda x: x[c.loc[x.index,"status"]=="failed"].sum()),
                                   fc=("status",lambda x: (x=="failed").sum()))
    pivot["total_value"] = agg["tv"]; pivot["failed_value"] = agg["fv"]
    pivot["successful_value"] = pivot["total_value"] - pivot["failed_value"]
    okc = pivot[[c for c in pivot.columns if c.endswith("_successful")]].sum(axis=1)
    pivot["success_rate"] = (okc / (okc + agg["fc"]).replace(0, pd.NA) * 100).round(1)
    pivot = pivot.reset_index()
    pivot.insert(1, "session_label", pivot["session"].map(lambda x: SESSION_LABEL_MAP.get(x,x)))
    meta = ["session","session_label"]; types = sorted(c for c in pivot.columns if "_successful"in c or "_failed"in c)
    vals = ["txns","successful_value","failed_value","total_value","success_rate"]
    exist = [c for c in meta+types+vals if c in pivot.columns]
    return pivot[exist]


def _compute_bank_summary(combined: pd.DataFrame) -> pd.DataFrame:
    """Bank performance from beneficiary_bvn (short codes)."""
    bvn = "beneficiary_bvn"
    if bvn not in combined.columns:
        return pd.DataFrame()
    gb = combined.groupby(bvn)
    okc = gb["status"].apply(lambda x: (x=="successful").sum())
    fc = gb["status"].apply(lambda x: (x=="failed").sum())
    tot = okc + fc
    has_amt = combined["amount"].notna()
    vg = combined[has_amt].groupby(bvn) if has_amt.any() else None
    tv = vg["amount"].sum() if vg is not None else pd.Series(dtype=float)
    fv = vg["amount"].apply(lambda x: x[combined.loc[x.index,"status"]=="failed"].sum()) if vg is not None else pd.Series(dtype=float)

    r = pd.DataFrame({"txns": tot, "ok": okc, "fail": fc,
                      "rate": (okc / tot.replace(0,pd.NA)*100).round(1)})
    r["total_value"] = tv.round(2); r["failed_value"] = fv.round(2); r["avg"] = (tv / tot.replace(0,pd.NA)).round(2)
    r = r.fillna(pd.NA).sort_values("txns", ascending=False).reset_index()
    r["bank_name"] = r[bvn].map(_bank_name)
    return r[["bank_name", bvn, "txns", "rate", "total_value", "failed_value", "avg"]]


def _compute_failure_analysis(combined: pd.DataFrame) -> pd.DataFrame:
    """Failure breakdown by NIP response code."""
    fail = combined[combined["status"]=="failed"].copy()
    if fail.empty: return pd.DataFrame()
    fail["code"] = fail.get("nip_response_code", pd.NA).fillna("N/A")
    gb = fail.groupby("code")
    fv = gb["amount"].apply(lambda x: x.fillna(0).sum())
    tot_fv = fv.sum()
    r = pd.DataFrame({"desc": gb.size().index.map(lambda c: NIP_CODE_MAP.get(c, "Name enquiry")),
                       "count": gb.size().values, "failed_value": fv.round(2).values,
                       "pct": (fv/tot_fv*100).round(1).values if tot_fv else 0})
    return r.sort_values("count", ascending=False).reset_index(drop=True)


def _compute_patterns(combined: pd.DataFrame) -> pd.DataFrame:
    """Narration category patterns."""
    if "narration" not in combined.columns: return pd.DataFrame()
    n = combined["narration"].dropna()
    if n.empty: return pd.DataFrame()
    cats = n.str.extract(r"^([^/\d]+)", expand=False).str.strip().str.lower()
    cats = cats.where(cats.notna() & (cats.str.len()>0), "other")
    c2 = combined.copy(); c2["cat"] = cats
    gb = c2.groupby("cat")
    ok = gb["status"].apply(lambda x: (x=="successful").sum()); tot = gb.size()
    tv = gb["amount"].sum().fillna(0)
    r = pd.DataFrame({"txns": tot, "ok": ok, "fail": tot-ok,
                       "rate": (ok/tot.replace(0,pd.NA)*100).round(1),
                       "total_value": tv.round(2)}).fillna(0)
    return r.sort_values("txns", ascending=False).head(8).reset_index()


def _compute_channel_summary(combined: pd.DataFrame) -> pd.DataFrame:
    """Channel breakdown."""
    if "channel_code" not in combined.columns: return pd.DataFrame()
    gb = combined.groupby("channel_code")
    ok = gb["status"].apply(lambda x: (x=="successful").sum()); tot = gb.size()
    r = pd.DataFrame({"code": gb.size().index, "txns": tot.values, "ok": ok.values,
                       "rate": (ok/tot.replace(0,pd.NA)*100).round(1).values})
    has_amt = combined["amount"].notna()
    if has_amt.any():
        tv = combined[has_amt].groupby("channel_code")["amount"].sum()
        r["total_value"] = r["code"].map(lambda c: round(tv.get(str(c), pd.NA), 2) if pd.notna(tv.get(str(c))) else pd.NA)
    else:
        r["total_value"] = pd.NA
    r["desc"] = r["code"].map(lambda x: CHANNEL_DESC.get(str(x), ""))
    return r.sort_values("txns", ascending=False).head(8).reset_index(drop=True)


def _write_header(ws, row: int, title: str, ncols: int) -> int:
    """Write a coloured section header across *ncols* columns.  Returns next row."""
    for c in range(1, ncols + 1):
        cl = ws.cell(row=row, column=c)
        cl.fill = HEADER_FILL; cl.border = THIN_BORDER
    ws.cell(row=row, column=1, value=title.upper()).font = Font(bold=True, size=11, color="FFFFFF")
    return row + 1


def _write_table(ws, row: int, df: pd.DataFrame, curr_cols: set | None = None) -> int:
    """Write a small table starting at *row*. Returns next available row."""
    if df.empty: return row
    curr_cols = curr_cols or set()
    for c, name in enumerate(df.columns, 1):
        cl = ws.cell(row=row, column=c, value=name)
        cl.fill = HEADER_FILL; cl.font = HEADER_FONT; cl.border = THIN_BORDER
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for _, dr in df.iterrows():
        row += 1
        for c, val in enumerate(dr, 1):
            cl = ws.cell(row=row, column=c, value=val if pd.notna(val) else None)
            cl.border = THIN_BORDER
            col_name = df.columns[c - 1]
            if col_name in curr_cols and cl.value is not None:
                cl.number_format = '"₦"#,##0.00'
    return row + 1


def _kpi_bar(ws, row: int, kpis: pd.DataFrame):
    """Write a compact KPI row with coloured cards."""
    if kpis.empty: return row + 1
    colors = [
        ("4472C4", "FFFFFF"), ("5B9BD5", "FFFFFF"), ("FF6347", "FFFFFF"),
        ("FF8C00", "FFFFFF"), ("70AD47", "FFFFFF"), ("C00000", "FFFFFF"),
        ("7030A0", "FFFFFF"), ("00B050", "FFFFFF"), ("FFC000", "000000"),
    ]
    for c, col_name in enumerate(kpis.columns, 1):
        val = kpis.iloc[0, c - 1]
        bg, fg = colors[c - 1] if c - 1 < len(colors) else ("D9E2F3", "000000")
        cl = ws.cell(row=row, column=c, value=col_name)
        cl.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cl.font = Font(bold=True, color=fg, size=9)
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cl.border = THIN_BORDER
        cl2 = ws.cell(row=row + 1, column=c, value=val)
        cl2.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cl2.font = Font(bold=True, size=12, color=bg)
        cl2.alignment = Alignment(horizontal="center", vertical="center")
        cl2.border = THIN_BORDER
        if "value" in col_name.lower() or "risk" in col_name.lower() or "avg" in col_name.lower():
            if isinstance(val, (int, float)):
                cl2.number_format = '"₦"#,##0.00'
    return row + 2



def _build_summary(ws, data: dict) -> int:
    """Dashboard with logo, KPI bar, charts, and 5 compact sections."""
    import os
    from openpyxl.drawing.image import Image

    parts = []
    for key in ("easy_pay", "direct_debit"):
        df = data.get(key)
        if df is not None and not df.empty:
            cols = [c for c in ["session","transaction_type","status","amount",
                                "destination_institution_code","channel_code",
                                "narration","nip_response_code","beneficiary_bvn"]
                    if c in df.columns]
            parts.append(df[cols])
    if not parts:
        ws.cell(row=1, column=1, value="(no data)"); return 0
    combined = pd.concat(parts, ignore_index=True)

    CURR = {"total_value","failed_value","successful_value","Total Value","At Risk",
            "Avg Value","total_value","failed_value","avg"}

    # ── Logo + Title ──
    logo_path = os.path.join(os.path.dirname(__file__), "..", "logo.png")
    if os.path.isfile(logo_path):
        img = Image(logo_path)
        img.width = 100; img.height = 100
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 70
    title_cell = ws.cell(row=1, column=3, value="NIBSS RECONCILIATION DASHBOARD")
    title_cell.font = Font(bold=True, size=18, color="1F4E79")
    date_str = datetime.now().strftime("%d %B %Y")
    ws.cell(row=2, column=3, value=f"Generated: {date_str}").font = Font(size=10, color="808080")
    row = 4

    # ── 1  KPI BAR ──
    row = _kpi_bar(ws, row, _compute_kpis(combined))

    # ── 2  SESSION OVERVIEW ──
    row += 1; row = _write_header(ws, row, "SESSION OVERVIEW", 12)
    row = _write_table(ws, row, _compute_session_overview(combined), CURR)

    # ── 3  BANK PERFORMANCE (cols A–G) + PIE CHART (cols I–N) ──
    row += 1
    bank_hdr = _write_header(ws, row, "TOP BANKS (by volume)", 7)
    bd = _compute_bank_summary(combined)
    row = _write_table(ws, bank_hdr, bd, CURR)
    bank_data_end = row

    # Failure analysis table in cols I–L (right of bank table)
    fd = _compute_failure_analysis(combined)
    if not fd.empty:
        fd_renamed = fd.rename(columns={"desc": "reason", "pct": "%"})
        fd_cols = [c for c in ["reason", "count", "failed_value", "%"] if c in fd_renamed.columns]
        # Section header at col I
        for c in range(9, 13):
            cl = ws.cell(row=bank_hdr - 1, column=c)
            cl.fill = HEADER_FILL; cl.border = THIN_BORDER
        ws.cell(row=bank_hdr - 1, column=9, value="FAILURES BY REASON").font = Font(bold=True, size=11, color="FFFFFF")
        # Column headers
        for ci, name in enumerate(fd_cols, 9):
            cl = ws.cell(row=bank_hdr, column=ci, value=name)
            cl.fill = HEADER_FILL; cl.font = HEADER_FONT; cl.border = THIN_BORDER
            cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Data rows
        for j, (_, dr) in enumerate(fd_renamed[fd_cols].iterrows()):
            r = bank_hdr + 1 + j
            for ci, val in enumerate(dr, 9):
                cl = ws.cell(row=r, column=ci, value=val if pd.notna(val) else None)
                cl.border = THIN_BORDER
                col_name = fd_cols[ci - 9]
                if col_name == "failed_value" and cl.value is not None:
                    cl.number_format = '"₦"#,##0.00'
                if col_name == "%" and cl.value is not None:
                    cl.number_format = '0.0"%"'

    # ── 4  TRANSACTION PATTERNS ──
    row = bank_data_end + 2
    row = _write_header(ws, row, "PURPOSE PATTERNS", 6)
    row = _write_table(ws, row, _compute_patterns(combined), CURR)

    # ── 5  CHANNEL BREAKDOWN ──
    row = _write_header(ws, row, "CHANNEL BREAKDOWN", 6)
    row = _write_table(ws, row, _compute_channel_summary(combined), CURR)

    # ── Footer note ──
    ws.cell(row=row + 1, column=1, value="Channel '2' = NIBSS NIP (standard instant payment channel). "
            "Other codes = originating institution IDs from name enquiries.").font = Font(italic=True, size=9, color="808080")

    return int(combined["session"].nunique())


def _build_data_sheet(ws, df: pd.DataFrame, table_name: str, run_log: list):
    """
    Sheet 2/3: Write a combined data table with AutoFilter, reordered
    columns, dead-column removal, currency formatting, and freeze panes.
    Rows where ``status == "failed"`` get a red background.
    """
    df = _reorder_columns(df)
    df = _remove_dead_columns(df, table_name.replace("Table", ""), run_log)

    _df_to_table(ws, df, table_name)

    if not df.empty:
        # Red fill on failed rows
        status_col = _find_col_letter(ws, "status")
        if status_col:
            _apply_red_to_failed(ws, status_col, len(df) + 1)

        # Currency format on amount
        amt_col = _find_col_letter(ws, "amount")
        if amt_col:
            _apply_currency_format(ws, amt_col, len(df) + 1)

        # Freeze row 1 and columns A–C (session, transaction_type, status)
        ws.freeze_panes = "D2"

    _apply_header_style(ws, len(df.columns) if not df.empty else 1)


def _build_exceptions(ws, data: dict, run_log: list):
    """
    Sheet 4: Every failed row from both sources, combined into one
    table with an added ``source`` column, sorted by amount descending.
    """
    ex_parts = []
    for key, label in [("easy_pay", "EasyPay"), ("direct_debit", "DirectDebit")]:
        df = data.get(key)
        if df is not None and not df.empty and "status" in df.columns:
            failed = df[df["status"] == "failed"].copy()
            if not failed.empty:
                failed["source"] = label
                ex_parts.append(failed)

    if not ex_parts:
        ws.cell(row=1, column=1, value="(no exceptions)")
        return

    combined = pd.concat(ex_parts, ignore_index=True, sort=False)

    # Reorder and drop dead columns before sorting (sort is on amount)
    combined = _reorder_columns(combined)
    combined = _remove_dead_columns(combined, "Exceptions", run_log)

    if "amount" in combined.columns:
        combined = combined.sort_values("amount", ascending=False, na_position="last")
    else:
        combined = combined.sort_values(["session", "transaction_type"])

    combined = combined.reset_index(drop=True)

    _df_to_table(ws, combined, "ExceptionsTable")

    if not combined.empty:
        status_col = _find_col_letter(ws, "status")
        if status_col:
            _apply_red_to_failed(ws, status_col, len(combined) + 1)

        amt_col = _find_col_letter(ws, "amount")
        if amt_col:
            _apply_currency_format(ws, amt_col, len(combined) + 1)

        ws.freeze_panes = "D2"

    _apply_header_style(ws, len(combined.columns))


def _build_run_log(ws, run_log: list, root_path: str, row_counts: dict):
    """
    Sheet 5: Metadata — generation timestamp, root folder, per-file
    manifest with file sizes / timestamps, and final row counts per
    data sheet.
    """
    ws.cell(row=1, column=1, value="Pipeline Run Log").font = Font(bold=True, size=14)

    ws.cell(row=2, column=1, value="Generated:")
    ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    ws.cell(row=3, column=1, value="Source Root:")
    ws.cell(row=3, column=2, value=str(root_path))

    # ── Row counts per sheet ──
    ws.cell(row=5, column=1, value="Sheet Row Counts").font = Font(bold=True, size=12)
    row = 6
    for sheet, count in row_counts.items():
        ws.cell(row=row, column=1, value=sheet)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # ── File manifest table ──
    row += 1
    ws.cell(row=row, column=1, value="File Manifest").font = Font(bold=True, size=12)
    row += 1
    headers = ["Session", "Source", "File", "Found", "Rows", "Size (KB)", "Modified", "Detail"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_style(ws, len(headers))
    row += 1

    for entry in run_log:
        ws.cell(row=row, column=1, value=entry.get("session", ""))
        ws.cell(row=row, column=2, value=entry.get("source", ""))
        ws.cell(row=row, column=3, value=entry.get("file", ""))
        ws.cell(row=row, column=4, value="Yes" if entry.get("found") else "No")
        ws.cell(row=row, column=5, value=entry.get("rows", 0))
        ws.cell(row=row, column=6, value=entry.get("size_kb", ""))
        ws.cell(row=row, column=7, value=entry.get("modified", ""))
        ws.cell(row=row, column=8, value=entry.get("detail", ""))
        row += 1


# ── Orchestrator ─────────────────────────────────────────────────────────────

def build_workbook(data: dict, run_log: list, root_path: str, output_path: str) -> str:
    """
    Create a 5-sheet Excel workbook at *output_path* and return the path.

    Parameters
    ----------
    data : dict
        Keys ``"easy_pay"`` and ``"direct_debit"`` mapping to DataFrames.
    run_log : list[dict]
        Per-file manifest from ``reader.load_all_data()``.
    root_path : str
        Original source folder (written into Run_Log).
    output_path : str
        Full destination path for the ``.xlsx`` file.
    """
    wb = Workbook()

    # Sheet 1 – Summary
    ws = wb.active
    ws.title = "Summary"
    n_sessions = _build_summary(ws, data)
    _auto_width(ws)

    # Sheet 2 – EasyPay
    ws = wb.create_sheet("EasyPay")
    _build_data_sheet(ws, data.get("easy_pay", pd.DataFrame()), "EasyPayTable", run_log)
    _auto_width(ws)

    # Sheet 3 – DirectDebit
    ws = wb.create_sheet("DirectDebit")
    _build_data_sheet(ws, data.get("direct_debit", pd.DataFrame()), "DirectDebitTable", run_log)
    _auto_width(ws)

    # Sheet 4 – Exceptions
    ws = wb.create_sheet("Exceptions")
    _build_exceptions(ws, data, run_log)
    _auto_width(ws)

    # Collect final row-counts for Run_Log
    row_counts = {
        "Summary Sessions": n_sessions,
        "EasyPay Rows": len(data.get("easy_pay", pd.DataFrame())),
        "DirectDebit Rows": len(data.get("direct_debit", pd.DataFrame())),
        "Exceptions Rows": _sheet_data_rows(wb["Exceptions"]),
    }

    # Sheet 5 – Run_Log
    ws = wb.create_sheet("Run_Log")
    _build_run_log(ws, run_log, root_path, row_counts)
    _auto_width(ws)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _sheet_data_rows(ws) -> int:
    """
    Return the number of data rows in a worksheet (total rows minus header),
    handling the "(no data)" placeholder case.
    """
    if ws.max_row <= 1:
        return 0
    first = ws.cell(row=1, column=1).value
    if first and str(first).startswith("(no "):
        return 0
    return ws.max_row - 1
